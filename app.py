import os
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from openpyxl import Workbook, load_workbook
import logging
import threading
import json

# Configuração do logging
logging.basicConfig(filename="app.log", level=logging.INFO, format="%(asctime)s - %(message)s")

# Nome padrão da planilha
dados_planilha = "dados_preenchidos.xlsx"
coordenadas_planilha = "Cadastro RAAG.xlsx"

# Arquivo para armazenar o nome da última planilha usada
CONFIG_FILE = "config.json"
json_file = "dados.json"

# Cache para armazenar os dados de coordenadas
coordenadas_cache = {}
classificacao_cache = {}

# Função para carregar ou criar a planilha
def carregar_ou_criar_planilha():
    """
    Carrega uma planilha existente ou cria uma nova se não existir.
    """
    global dados_planilha

    # Verifica se o arquivo de configuração existe
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            config = json.load(f)
            dados_planilha = config.get("ultima_planilha", dados_planilha)  # Obtém o nome salvo, se existir
            
    else:
        # Se não existir, solicita o nome da planilha ao usuário
        nome_planilha = simpledialog.askstring("Nome da Planilha", "Digite o nome da planilha:")
        if nome_planilha:
            dados_planilha = nome_planilha + ".xlsx"
            with open(CONFIG_FILE, "w") as f:
                json.dump({"ultima_planilha": dados_planilha}, f)  # Salva o nome no JSON
        else:
            messagebox.showerror("Erro", "Nenhum nome foi fornecido. O programa será encerrado.")
            root.destroy()
            return

    # Verifica se a planilha já existe
    if not os.path.exists(dados_planilha):
        # Cria uma nova planilha com a aba "Levantamento"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Levantamento"
        # Adiciona os cabeçalhos
        cabecalhos = [
            "ID RAAG", "ID IPPUC", "Via", "Bairro", "Classificação", "Distribuição", "Coordenadas",
            "Largura do Passeio Adjacente", "Largura do Gramado Adjacente", "Largura do Estacionamento Adjacente", 
            "Largura da Pista 1", "Largura do Canteiro Central", "Largura da Pista 2", 
            "Largura do Estacionamento Oposto", "Largura do Gramado Oposto", "Largura do Passeio Oposto", "Ciclovia",
            "Distância entre Postes", "Altura", "Projeção", "Recuo", "Interferência Arbórea", "Led"
        ]
        sheet.append(cabecalhos)
        workbook.save(dados_planilha)
        logging.info(f"Nova planilha '{dados_planilha}' criada com sucesso!")
    else:
        logging.info(f"Usando a planilha existente: {dados_planilha}")

    root.attributes('-topmost', False)  # Remove a janela do topo

# Função para carregar coordenadas na memória
def carregar_coordenadas_na_memoria():
    global coordenadas_cache
    try:
        if os.path.exists(coordenadas_planilha):
            workbook = load_workbook(coordenadas_planilha, read_only=True)
            sheet = workbook.active
            coordenadas_cache = {
                str(row[0]): {
                    "coordenadas": f"{row[7]},{row[8]}",
                    "bairro": row[5],
                    "distancia_postes": row[45],
                    "altura": row[39],
                    "projecao": row[38],
                    "recuo": row[43],
                }
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] is not None and row[7] is not None and row[8] is not None and row[45] is not None and row[39] is not None and row[38] is not None and row[43] is not None
            }
            logging.info("Coordenadas e bairros carregados na memória: %d registros.", len(coordenadas_cache))
        else:
            coordenadas_cache = {}
            logging.warning("Arquivo de coordenadas não encontrado.")
    except Exception as e:
        logging.error("Erro ao carregar coordenadas na memória: %s", e)
        coordenadas_cache = {}

# Função para buscar coordenadas
def buscar_coordenadas(id_raag):
    try:
        dados = coordenadas_cache.get(id_raag, {})
        return dados.get("coordenadas", ""), dados.get("bairro", ""), dados.get("distancia_postes", ""), dados.get("altura", ""), dados.get("projecao", ""), dados.get("recuo", "")
    except Exception as e:
        logging.error("Erro ao buscar coordenadas e bairro: %s", e)
        messagebox.showerror("Erro", "Ocorreu um erro ao buscar os dados.")
        return "", "", "", "", "", ""

# Função para preencher dados automaticamente
def preencher_dados_automaticamente(event=None):
    id_raag = entries["entry_id_raag"].get().strip()
    if id_raag:
        coordenadas, bairro, distancia_postes, altura, projecao, recuo = buscar_coordenadas(id_raag)

        if coordenadas:
            try:
                latitude, longitude = coordenadas.split(",")
                if not entries["entry_Coordenadas"].get().strip():  # Apenas preenche se estiver vazio
                    entries["entry_Coordenadas"].delete(0, tk.END)
                    entries["entry_Coordenadas"].insert(0, f"{latitude.strip()}, {longitude.strip()}")
            except ValueError:
                messagebox.showerror("Erro", "Coordenadas inválidas no banco de dados.")

        if bairro and not entries["entry_bairro"].get().strip():  # Apenas preenche se estiver vazio
            entries["entry_bairro"].delete(0, tk.END)
            entries["entry_bairro"].insert(0, bairro.strip())

        if distancia_postes and not entries["entry_distancia_postes"].get().strip():  # Apenas preenche se estiver vazio
            entries["entry_distancia_postes"].delete(0, tk.END)
            entries["entry_distancia_postes"].insert(0, str(distancia_postes))
        
        if altura and not entries["entry_altura"].get().strip():  # Apenas preenche se estiver vazio
            entries["entry_altura"].delete(0, tk.END)
            entries["entry_altura"].insert(0, str(altura))
        
        if projecao and not entries["entry_projecao"].get().strip():  # Apenas preenche se estiver vazio
            entries["entry_projecao"].delete(0, tk.END)
            entries["entry_projecao"].insert(0, str(projecao))
        
        if recuo and not entries["entry_recuo"].get().strip():  # Apenas preenche se estiver vazio
            entries["entry_recuo"].delete(0, tk.END)
            entries["entry_recuo"].insert(0, str(recuo))

# Função para carregar classificação na memória
def carregar_classificacao_na_memoria():
    global classificacao_cache
    try:
        if os.path.exists("Classificação.xlsx"):
            workbook = load_workbook("Classificação.xlsx", read_only=True)
            sheet = workbook.active
            classificacao_cache = {
                str(row[0]).strip().upper(): (row[1], row[2])  # Normaliza o nome da via
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] and row[1] and row[2]
            }
            logging.info("Classificação carregada na memória: %d registros.", len(classificacao_cache))
        else:
            classificacao_cache = {}
            logging.warning("Arquivo de classificação não encontrado.")
    except Exception as e:
        logging.error("Erro ao carregar classificação na memória: %s", e)
        classificacao_cache = {}

# Função para buscar classificação
def buscar_classificacao(via):
    try:
        # Normaliza o nome da via (converte para maiúsculas)
        via_normalizada = via.strip().upper()
        return classificacao_cache.get(via_normalizada, ("", ""))  # Retorna tupla com Classe Via e Passeio
    except Exception as e:
        logging.error("Erro ao buscar classificação: %s", e)
        messagebox.showerror("Erro", "Ocorreu um erro ao buscar a classificação.")
        return "", ""

# Função para preencher classificação
def preencher_classificacao(event):
    via = entries["entry_via"].get().strip()
    if via:
        # Normaliza o nome da via (converte para maiúsculas)
        via_normalizada = via.upper()
        classe_via, passeio = buscar_classificacao(via)
        if classe_via and passeio and not entries["entry_classificacao"].get().strip():  
            classificacao_completa = f"{classe_via}-{passeio}"
            entries["entry_classificacao"].delete(0, tk.END)
            entries["entry_classificacao"].insert(0, classificacao_completa)

# Função para salvar dados
def salvar_dados():
    if not validar_campos():
        return

    # Executa a função de salvar em segundo plano
    threading.Thread(target=salvar_em_segundo_plano, daemon=True).start()


# Função para exibir a mensagem temporária
def exibir_mensagem_temporaria(mensagem):
    label_mensagem.config(text=mensagem)  # Exibe a mensagem
    label_mensagem.grid()  # Torna o Label visível
    root.after(4000, lambda: label_mensagem.grid_remove())  # Oculta após 4 segundos

# Modifique a função salvar_em_segundo_plano
def salvar_em_segundo_plano():
    try:
        # Carrega a planilha
        workbook = load_workbook(dados_planilha)
        sheet = workbook["Levantamento"]

        # Coleta os dados do formulário
        dados = {key: entry.get().strip() for key, entry in entries.items()}

        id_raag = dados["entry_id_raag"]
        linha_existente = None

        # Verifica se o ID RAAG já existe na planilha
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == id_raag:
                linha_existente = row[0].row
                break

        # Campos que devem manter os valores atuais caso não sejam preenchidos novamente
        campos_preservados = {
            "entry_largura_passeio_adj", "entry_largura_gramado_adj", "entry_largura_estac_adj",
            "entry_largura_pista1", "entry_largura_canteiro_central", "entry_largura_pista2",
            "entry_largura_estac_opo", "entry_largura_gramado_opo", "entry_largura_passeio_opo",
            "entry_ciclovia", "entry_via", "entry_bairro", "entry_classificacao", "combobox_distribuicao"
        }

        if linha_existente:
            # Editar registro existente
            valores_atuais = {}  # Armazena valores atuais antes da edição
            for idx, key in enumerate(entries.keys(), start=1):
                valores_atuais[key] = sheet.cell(row=linha_existente, column=idx).value

            for idx, (chave, valor) in enumerate(dados.items(), start=1):
                if chave == "entry_id_raag" and valor.strip() != str(sheet.cell(row=linha_existente, column=1).value):
                    # Verifica se o novo ID RAAG é diferente do atual na planilha e se já existe em outra linha
                    if any(str(cell.value) == valor for row in sheet.iter_rows(min_row=2) for cell in row if cell.row != linha_existente):
                        root.after(0, messagebox.showerror, "Erro", "Este ID RAAG já existe na planilha.")
                        return  # Impede a edição se o novo ID RAAG já existir
                    else:
                        sheet.cell(row=linha_existente, column=idx, value=valor)  # Permite a alteração do ID RAAG
                elif chave in campos_preservados and not valor.strip():
                    # Mantém o valor atual para os outros campos preservados se estiverem vazios
                    sheet.cell(row=linha_existente, column=idx, value=valores_atuais[chave])
                else:
                    # Atualiza outros campos normalmente
                    sheet.cell(row=linha_existente, column=idx, value=valor)

            root.after(0, exibir_mensagem_temporaria, f"Dados do ID RAAG {id_raag} editados com sucesso!")

        else:  # Inserir novo registro
            prox_linha = sheet.max_row + 1
            for idx, valor in enumerate(dados.values(), start=1):
                sheet.cell(row=prox_linha, column=idx, value=valor)
            root.after(0, exibir_mensagem_temporaria, "Dados salvos com sucesso!")

        workbook.save(dados_planilha)
        logging.info("Dados salvos/editados: %s", dados)

        # Limpa os campos (exceto os preservados)
        for key, entry in entries.items():
            if key not in campos_preservados:
                if isinstance(entry, ttk.Combobox):
                    entry.set("Selecione")
                else:
                    entry.delete(0, tk.END)

        atualizar_treeview()

    except Exception as e:
        logging.error("Erro ao salvar dados: %s", e)
        root.after(0, messagebox.showerror, "Erro", f"Ocorreu um erro ao salvar os dados: {e}")

def atualizar_treeview():
    global vias
    vias = carregar_vias_e_ids()
    for item in tree_vias.get_children():
        tree_vias.delete(item)

    for via, ids in vias.items():
        # Converte os IDs para strings antes de inseri-los na Treeview
        ids_str = [str(id_raag) for id_raag in ids]  # ou tuple(map(str, ids))
        tree_vias.insert("", "end", text=via, values=ids_str)

    filtrar_vias()

# Função para validar campos
def validar_campos():
    for key, entry in entries.items():
        if isinstance(entry, ttk.Combobox) and entry.get() == "Selecione":
            messagebox.showerror("Erro", f"O campo '{key}' deve ser preenchido.")
            return False
        elif isinstance(entry, tk.Entry) and not entry.get().strip():
            # Permitir que apenas o campo ID IPPUC fique em branco
            if key == "entry_id_ippuc":
                continue
            messagebox.showerror("Erro", f"O campo '{key}' deve ser preenchido.")
            return False
    return True

# Função para limpar campos
def limpar_campos():
    resposta = messagebox.askyesno("Confirmação", "Tem certeza de que deseja limpar todos os campos?")
    if resposta:  # Apenas limpa os campos se a resposta for "Sim"
        for entry in entries.values():
            if isinstance(entry, ttk.Combobox):
                entry.set("Selecione")
            else:
                entry.delete(0, tk.END)

def deletar_id_raag(event=None):
    item_selecionado = tree_vias.selection()
    if item_selecionado:
        item = item_selecionado[0]
        valores = tree_vias.item(item, "values")
        if valores:  # Verifica se é um ID RAAG (tem valores)
            id_raag = valores[0]
            confirmacao = messagebox.askyesno("Confirmar Exclusão", f"Tem certeza que deseja excluir o ID RAAG {id_raag}?")
            if confirmacao:
                try:
                    workbook = load_workbook(dados_planilha)
                    sheet = workbook["Levantamento"]

                    linha_para_deletar = None
                    for row in sheet.iter_rows(min_row=2):
                        if str(row[0].value) == id_raag:
                            linha_para_deletar = row[0].row
                            break

                    if linha_para_deletar:
                        sheet.delete_rows(linha_para_deletar)
                        workbook.save(dados_planilha)

                        # Remove o ID RAAG da Treeview
                        tree_vias.delete(item)
                        atualizar_treeview()  # Atualiza a treeview após a exclusão

                        messagebox.showinfo("Sucesso", f"ID RAAG {id_raag} excluído com sucesso!")
                        limpar_campos()

                    else:
                        messagebox.showwarning("Aviso", f"ID RAAG {id_raag} não encontrado na planilha.")

                except Exception as e:
                    logging.error("Erro ao excluir ID RAAG: %s", e)
                    messagebox.showerror("Erro", f"Ocorreu um erro ao excluir o ID RAAG: {e}")

# Função para mudar de aba
def mudar_aba(direcao):
    atual = notebook.index(notebook.select())
    if direcao == "proxima" and atual < len(notebook.tabs()) - 1:
        notebook.select(atual + 1)
    elif direcao == "anterior" and atual > 0:
        notebook.select(atual - 1)

# Função para carregar vias e IDs RAAG
def carregar_vias_e_ids():
    try:
        workbook = load_workbook(dados_planilha)
        sheet = workbook["Levantamento"]
        vias = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] and row[0]:  # Via e ID RAAG
                via = row[2]
                id_raag = str(row[0])
                if via not in vias:
                    vias[via] = []
                vias[via].append(id_raag)
        return vias
    except Exception as e:
        logging.error(f"Erro ao carregar vias e IDs RAAG: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro ao carregar as vias e IDs RAAG: {e}")
        return {}

# Função para filtrar vias
def filtrar_vias(event=None):
    termo = entry_busca_via.get().strip().lower()
    for via in tree_vias.get_children():
        tree_vias.delete(via)
    for via, ids in vias.items():
        if termo in via.lower():
            tree_vias.insert("", "end", text=via, values=ids)

# Função para carregar dados por ID RAAG
def carregar_dados_por_id(id_raag):
    try:
        workbook = load_workbook(dados_planilha)
        sheet = workbook["Levantamento"]
        dados_encontrados = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == id_raag:
                dados_encontrados = row
                break

        if dados_encontrados:
            for idx, valor in enumerate(dados_encontrados):
                coluna = list(entries.keys())[idx]
                if valor:  # Apenas insere se houver valor
                    if isinstance(entries[coluna], ttk.Combobox):
                        entries[coluna].set(valor)
                    else:
                        entries[coluna].delete(0, tk.END)
                        entries[coluna].insert(0, valor)
            messagebox.showinfo("Sucesso", f"Dados do ID RAAG {id_raag} carregados com sucesso!")
        else:
            messagebox.showwarning("Aviso", f"ID RAAG {id_raag} não encontrado.")

    except Exception as e:
        logging.error("Erro ao carregar dados: %s", e)
        messagebox.showerror("Erro", "Ocorreu um erro ao carregar os dados.")

def selecionar_via_ou_id(event):
    item = tree_vias.selection()[0]
    texto = tree_vias.item(item, "text")
    valores = tree_vias.item(item, "values")

    if texto:  # Verifica se o texto não está vazio (é uma via)
        if tree_vias.item(item, 'open'):
            tree_vias.item(item, open=False)  # Recolhe a via se estiver aberta
        else:
            if texto in vias:
                # Verifica se a via já tem filhos (IDs RAAG)
                if not tree_vias.get_children(item):  # Insere somente se não houver filhos
                    for id_raag in vias[texto]:
                        tree_vias.insert(item, "end", text="", values=(str(id_raag),))
                tree_vias.item(item, open=True)  # Expande a via

    elif valores:  # Se não for uma via e tiver valores, é um ID RAAG
        id_raag = valores[0]
        carregar_dados_por_id(id_raag)

        # Expande a via pai se não estiver expandida (mesma lógica anterior)
        pai = tree_vias.parent(item)
        if pai and not tree_vias.item(pai, 'open'):
            tree_vias.item(pai, open=True)

# Função para recolher/expandir a seção lateral
def toggle_frame_vias():
    if frame_vias.winfo_ismapped():
        frame_vias.grid_remove()  # Use grid_remove()
        botao_toggle.config(text="Mostrar Vias")
    else:
        frame_vias.grid()      # Use grid()
        botao_toggle.config(text="Recolher Vias")
        root.update_idletasks() # Force layout update


# Inicialização da interface gráfica
root = tk.Tk()
carregar_ou_criar_planilha()
root.title("Preencher Dados na Planilha")
root.geometry("1000x600")
root.attributes('-topmost', True)


# Carregar coordenadas na memória
carregar_coordenadas_na_memoria()

# Carregar classificação na memória
carregar_classificacao_na_memoria()

# Frame fixo à esquerda para vias e busca
frame_vias = ttk.Frame(root)
frame_vias.grid(row=0, column=0, sticky="ns", padx=10, pady=10)

# Conteúdo do frame_vias
label_busca_via = ttk.Label(frame_vias, text="Buscar Via:")
label_busca_via.pack(pady=5)

entry_busca_via = ttk.Entry(frame_vias)
entry_busca_via.pack(pady=5, fill="x")
entry_busca_via.bind("<KeyRelease>", filtrar_vias)

# Configuração da Treeview para exibir vias e IDs RAAG
tree_vias = ttk.Treeview(frame_vias, columns=("ids_raag",), show="tree headings")
tree_vias.heading("#0", text="Via")
tree_vias.heading("ids_raag", text="IDs RAAG")
tree_vias.column("ids_raag", width=150)  # Define a largura da coluna
tree_vias.pack(fill="both", expand=True)
tree_vias.bind("<Double-1>", selecionar_via_ou_id)

# Carregar vias e IDs RAAG
vias = carregar_vias_e_ids()
for via, ids in vias.items():
    tree_vias.insert("", "end", text=via)

# Interface gráfica principal (Corrected to use grid)
notebook = ttk.Notebook(root)
notebook.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)

aba1 = ttk.Frame(notebook)
aba2 = ttk.Frame(notebook)
aba3 = ttk.Frame(notebook)
notebook.add(aba1, text="Dados Gerais")
notebook.add(aba2, text="Larguras e Ciclovia")
notebook.add(aba3, text="Postes e Interferências")

for aba in [aba1, aba2, aba3]:
    aba.grid_columnconfigure(0, weight=1)
    aba.grid_columnconfigure(1, weight=2)

# Função para criar campos na interface
def criar_campos(aba, fields):
    for idx, (label_text, field_name) in enumerate(fields, start=1):
        label = ttk.Label(aba, text=label_text)
        label.grid(row=idx, column=0, padx=5, pady=5, sticky=tk.W)

        if "combobox" in field_name.lower():
            values = ["Selecione"]
            if field_name == "combobox_distribuicao":
                values.extend(["UNILATERAL", "BILATERAL ALTERNADO", "BILATERAL FF"])
            elif field_name == "combobox_interferencia_arborea":
                values.extend(["SIM", "NÃO"])
            elif field_name == "combobox_led":
                values.extend(["SIM", "NÃO"])

            combobox = ttk.Combobox(aba, values=values, state="readonly")
            combobox.set("Selecione")
            combobox.grid(row=idx, column=1, padx=5, pady=5, sticky="ew")
            entries[field_name] = combobox
        else:
            entry = ttk.Entry(aba)
            entry.grid(row=idx, column=1, padx=5, pady=5, sticky="ew")
            entries[field_name] = entry

# Dicionário para armazenar os campos
entries = {}

# Criar campos na aba 1
criar_campos(aba1, [
    ("ID RAAG", "entry_id_raag"),
    ("ID IPPUC", "entry_id_ippuc"),
    ("Via", "entry_via"),
    ("Bairro", "entry_bairro"),
    ("Classificação", "entry_classificacao"),
    ("Distribuição", "combobox_distribuicao"),
    ("Coordenadas", "entry_Coordenadas"),
])

# Associar evento para preenchimento automático da classificação
entries["entry_via"].bind("<FocusOut>", preencher_classificacao)

# Criar campos na aba 2
criar_campos(aba2, [
    ("Largura do Passeio Adjacente", "entry_largura_passeio_adj"),
    ("Largura do Gramado Adjacente", "entry_largura_gramado_adj"),
    ("Largura do Estacionamento Adjacente", "entry_largura_estac_adj"),
    ("Largura da Pista 1", "entry_largura_pista1"),
    ("Largura do Canteiro Central", "entry_largura_canteiro_central"),
    ("Largura da Pista 2", "entry_largura_pista2"),
    ("Largura do Estacionamento Oposto", "entry_largura_estac_opo"),
    ("Largura do Gramado Oposto", "entry_largura_gramado_opo"),
    ("Largura do Passeio Oposto", "entry_largura_passeio_opo"),
    ("Ciclovia", "entry_ciclovia"),
])

# Criar campos na aba 3
criar_campos(aba3, [
    ("Distância entre Postes", "entry_distancia_postes"),
    ("Altura", "entry_altura"),
    ("Projeção", "entry_projecao"),
    ("Recuo", "entry_recuo"),
    ("Interferência Arbórea", "combobox_interferencia_arborea"),
    ("Led", "combobox_led"),
])


# Adicione um Label para a mensagem temporária
label_mensagem = ttk.Label(root, text="", foreground="green")
label_mensagem.grid(row=3, column=1, sticky="ew", padx=10, pady=10)

# Botões principais (incluindo o botão de alternar vias)
frame_botoes = ttk.Frame(root)
frame_botoes.grid(row=2, column=1, sticky="ew", padx=10, pady=10)

botao_salvar = ttk.Button(frame_botoes, text="Salvar", command=salvar_dados)
botao_salvar.grid(row=0, column=0, padx=5, pady=5)

botao_limpar = ttk.Button(frame_botoes, text="Limpar Campos", command=limpar_campos)
botao_limpar.grid(row=0, column=1, padx=5, pady=5)

botao_anterior = ttk.Button(frame_botoes, text="Aba Anterior", command=lambda: mudar_aba("anterior"))
botao_anterior.grid(row=0, column=2, padx=5, pady=5)

botao_proxima = ttk.Button(frame_botoes, text="Próxima Aba", command=lambda: mudar_aba("proxima"))
botao_proxima.grid(row=0, column=3, padx=5, pady=5)

botao_toggle = ttk.Button(frame_botoes, text="Recolher Vias", command=toggle_frame_vias)
botao_toggle.grid(row=0, column=4, padx=5, pady=5)  # Correctly place

tree_vias.bind("<Double-1>", selecionar_via_ou_id)

entries["entry_id_raag"].bind("<FocusOut>", preencher_dados_automaticamente)

atualizar_treeview()  # Inicializa a Treeview

tree_vias.bind("<Delete>", deletar_id_raag)

# Iniciar a interface gráfica
root.mainloop()


