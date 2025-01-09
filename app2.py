import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
import logging
import shutil

# Configuração do logging
logging.basicConfig(filename="app.log", level=logging.INFO, format="%(asctime)s - %(message)s")

def selecionar_planilha():
    global caminho_planilha
    caminho_planilha = filedialog.askopenfilename(title="Selecione a planilha", filetypes=[("Excel files", "*.xlsm")])
    if not caminho_planilha:
        messagebox.showerror("Erro", "Nenhuma planilha selecionada!")
        return False
    return True

def validar_dados():
    try:
        float(entry_largura_passeio_adj.get())  # Exemplo de validação numérica
        float(entry_largura_passeio_opo.get())
        # Adicione validações conforme necessário
    except ValueError:
        messagebox.showerror("Erro", "Certifique-se de que os campos numéricos estão preenchidos corretamente.")
        return False
    return True

def salvar_dados():
    if not validar_dados():
        return

    # Backup da planilha
    backup_path = caminho_planilha.replace(".xlsm", "_backup.xlsm")
    shutil.copyfile(caminho_planilha, backup_path)
    logging.info("Backup criado: %s", backup_path)

    # Obtendo os valores dos campos
    dados = {
        "ID RAAG": entry_id_raag.get(),
        "Via": entry_via.get(),
        "Classificação": entry_classificacao.get(),
        "Distribuição": entry_distribuicao.get(),
        "Largura do Passeio Adjacente": entry_largura_passeio_adj.get(),
        "Largura do Passeio Oposto": entry_largura_passeio_opo.get(),
        "Largura do Gramado Adjacente": entry_largura_gramado_adj.get(),
        "Largura do Gramado Oposto": entry_largura_gramado_opo.get(),
        "Largura do Estacionamento Adjacente": entry_largura_estac_adj.get(),
        "Largura do Estacionamento Oposto": entry_largura_estac_opo.get(),
        "Largura da Pista 1": entry_largura_pista1.get(),
        "Largura do Canteiro Central": entry_largura_canteiro_central.get(),
        "Largura da Pista 2": entry_largura_pista2.get(),
        "Ciclovia": entry_ciclovia.get(),
        "Distância entre Postes": entry_distancia_postes.get(),
        "Altura": entry_altura.get(),
        "Projeção": entry_projecao.get(),
        "Interferência Arbórea": entry_interferencia_arborea.get(),
    }

    # Carregar a planilha
    workbook = load_workbook(caminho_planilha, keep_vba=True)
    sheet = workbook['Levantamento']

    # Encontrar a próxima linha vazia
    prox_linha = sheet.max_row + 1

    # Preencher os dados na planilha
    colunas = list(dados.keys())
    for idx, valor in enumerate(dados.values(), start=1):
        sheet.cell(row=prox_linha, column=idx, value=valor)

    # Salvar a planilha
    workbook.save(caminho_planilha)
    logging.info("Dados salvos: %s", dados)

    # Mensagem de confirmação
    messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")

def limpar_campos():
    for entry in entries.values():
        entry.delete(0, tk.END)

# Criar a interface gráfica
root = tk.Tk()
root.title("Preencher Dados na Planilha")

# Botão para selecionar a planilha
button_selecionar = ttk.Button(root, text="Selecionar Planilha", command=selecionar_planilha)
button_selecionar.grid(row=0, column=0, columnspan=2, pady=10)

# Campos do formulário
fields = [
    ("ID RAAG", "entry_id_raag"),
    ("Via", "entry_via"),
    ("Classificação", "entry_classificacao"),
    ("Distribuição", "entry_distribuicao"),
    ("Largura do Passeio Adjacente", "entry_largura_passeio_adj"),
    ("Largura do Passeio Oposto", "entry_largura_passeio_opo"),
    ("Largura do Gramado Adjacente", "entry_largura_gramado_adj"),
    ("Largura do Gramado Oposto", "entry_largura_gramado_opo"),
    ("Largura do Estacionamento Adjacente", "entry_largura_estac_adj"),
    ("Largura do Estacionamento Oposto", "entry_largura_estac_opo"),
    ("Largura da Pista 1", "entry_largura_pista1"),
    ("Largura do Canteiro Central", "entry_largura_canteiro_central"),
    ("Largura da Pista 2", "entry_largura_pista2"),
    ("Ciclovia", "entry_ciclovia"),
    ("Distância entre Postes", "entry_distancia_postes"),
    ("Altura", "entry_altura"),
    ("Projeção", "entry_projecao"),
    ("Interferência Arbórea", "entry_interferencia_arborea"),
]

entries = {}
for idx, (label_text, entry_var) in enumerate(fields, start=1):
    label = ttk.Label(root, text=label_text)
    label.grid(row=idx, column=0, padx=5, pady=5, sticky=tk.W)
    entry = ttk.Entry(root)
    entry.grid(row=idx, column=1, padx=5, pady=5, sticky=tk.E)
    entries[entry_var] = entry

globals().update(entries)

# Botões de ação
button_salvar = ttk.Button(root, text="Salvar Dados", command=salvar_dados)
button_salvar.grid(row=len(fields) + 1, column=0, pady=10)

button_limpar = ttk.Button(root, text="Limpar Campos", command=limpar_campos)
button_limpar.grid(row=len(fields) + 1, column=1, pady=10)

root.mainloop()
