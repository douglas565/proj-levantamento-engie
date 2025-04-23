# Create by Douglas Ramos Charqueiro
# Developed in ESIP


import os
import openpyxl
import logging
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from openpyxl import Workbook, load_workbook
import threading
import json
import time
import csv
import sys


# Inicialização da interface gráfica
root = tk.Tk()
root.attributes("-topmost", True)  # Mantém a janela sempre no topo
root.title("Levantamento de Iluminação Pública")
root.withdraw()  # Esconde a janela principal inicialmente


class LoadingScreen:
    def __init__(self, parent):
        self.parent = parent
        self.window = tk.Toplevel(parent)
        self.window.title("Carregando...")
        self.window.geometry("300x150")
        self.window.resizable(False, False)
        
        # Centraliza a janela
        window_width = 300
        window_height = 150
        position_right = int(self.window.winfo_screenwidth()/2 - window_width/2)
        position_down = int(self.window.winfo_screenheight()/2 - window_height/2)
        self.window.geometry(f"+{position_right}+{position_down}")
        
        # Remove a decoração da janela
        self.window.overrideredirect(True)
        
        # Frame principal
        frame = ttk.Frame(self.window)
        frame.pack(expand=True, fill='both', padx=20, pady=20)
        
        # Label de carregamento
        self.label = ttk.Label(frame, text="Iniciando sistema...", font=('Arial', 10))
        self.label.pack(pady=10)
        
        # Barra de progresso (modo determinístico para controle preciso)
        self.progress = ttk.Progressbar(frame, mode='determinate', maximum=100)
        self.progress.pack(fill='x', pady=10)
        
        # Inicia a barra de progresso
        self.progress["value"] = 0
        self.window.update()  # Força a atualização da tela

    def update_progress(self, value, text):
        """Atualiza a barra de progresso e o texto de forma assíncrona."""
        self.progress["value"] = value
        self.label.config(text=text)
        self.window.update_idletasks()  # Atualiza a interface sem travar

    def close(self):
        self.window.destroy()

def show_credits():
    about_window = tk.Toplevel()
    about_window.title("Sobre o Sistema")
    about_window.geometry("400x300")
    
    # Frame principal
    frame_principal = ttk.Frame(about_window)
    frame_principal.pack(expand=True, fill='both', padx=20, pady=20)
    
    try:
        # Tenta carregar a logo
        logo_path = os.path.join(os.path.dirname(__file__), "engie_logo.png")
        if os.path.exists(logo_path):
            logo = tk.PhotoImage(file=logo_path)
            logo_label = ttk.Label(frame_principal, image=logo)
            logo_label.image = logo
            logo_label.pack(pady=10)
    except Exception as e:
        logging.error(f"Erro ao carregar logo: {e}")
    
    # Texto dos créditos
    credits = ttk.Label(frame_principal, text="""ENGIE SOLUÇÕES
Sistema de Gestão de Iluminação Pública

Versão 1.3.0

Developed by: ESIP
Created by: Douglas Ramos Charqueiro

© 2025 Todos os direitos reservados""", justify='center')
    credits.pack(pady=10)
    
    # Botão de fechar
    ttk.Button(frame_principal, text="Fechar", 
              command=about_window.destroy).pack(pady=10)
    


# Configuração do logging
logging.basicConfig(filename="app.log", level=logging.INFO, format="%(asctime)s - %(message)s")

# Nome padrão da planilha
dados_planilha = "dados_preenchidos.xlsx"
coordenadas_planilha = "Cadastro RAAG.xlsx"
coordenadas_ippuc_planilha = "id_ippuc_coordenadas.xlsx"

# Arquivo para armazenar o nome da última planilha usada
CONFIG_FILE = "config.json"
json_file = "dados.json"

# Cache para armazenar os dados de coordenadas
coordenadas_cache = {}
classificacao_cache = {}
ippuc_cache = {}

# Função para carregar ou criar a planilha
def carregar_ou_criar_planilha():
    """
    Carrega uma planilha existente ou cria uma nova se não existir.
    Retorna True se a planilha está pronta para uso, False caso contrário.
    """
    global dados_planilha
    
    try:
        # Verifica se existe arquivo de configuração
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r") as f:
                config = json.load(f)
                dados_planilha = config.get("ultima_planilha", "dados_preenchidos.xlsx")
                logging.info(f"Configuração carregada. Planilha: {dados_planilha}")
        else:
            # PRIMEIRA EXECUÇÃO - PEDE NOME DA PLANILHA
            root.deiconify()  # Mostra a janela principal para poder exibir o diálogo
            nome_planilha = simpledialog.askstring(
                "Nome da Planilha", 
                "Digite o nome para a nova planilha (sem extensão):",
                parent=root
            )
            
            if not nome_planilha or nome_planilha.strip() == "":
                messagebox.showerror("Erro", "É necessário fornecer um nome para a planilha.")
                return False
                
            dados_planilha = f"{nome_planilha.strip()}.xlsx"
            
            # Salva no arquivo de configuração
            with open(CONFIG_FILE, "w") as f:
                json.dump({"ultima_planilha": dados_planilha}, f)
            
            logging.info(f"Nova planilha configurada: {dados_planilha}")
            root.withdraw()  # Volta a esconder a janela principal

        # Se a planilha não existe, cria uma nova
        if not os.path.exists(dados_planilha):
            try:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Levantamento"
                
                # Cabeçalhos completos
                cabecalhos = [
                    "ID RAAG", "ID IPPUC", "Via", "Bairro", "Trecho", "Classificação", "Distribuição",
                    "Latitude", "Longitude", "Largura do Passeio Adjacente", "Largura do Gramado Adjacente",
                    "Largura do Estacionamento Adjacente", "Largura da Pista 1", "Largura do Canteiro Central",
                    "Largura da Pista 2", "Largura do Estacionamento Oposto", "Largura do Gramado Oposto",
                    "Largura do Passeio Oposto", "Ciclovia", "Distância entre Postes", "Altura", "Projeção",
                    "Recuo", "Interferência Arbórea", "Led", "Modelo do LED", "Observações", "Observações Gerais"
                ]
                sheet.append(cabecalhos)
                
                workbook.save(dados_planilha)
                workbook.close()
                
                logging.info(f"Nova planilha criada: {os.path.abspath(dados_planilha)}")
                messagebox.showinfo("Sucesso", f"Planilha '{dados_planilha}' criada com sucesso!")
            except Exception as e:
                logging.error(f"Erro ao criar planilha: {e}")
                messagebox.showerror("Erro", f"Não foi possível criar a planilha: {e}")
                return False
        
        return True
        
    except Exception as e:
        logging.error(f"Erro crítico em carregar_ou_criar_planilha: {e}")
        messagebox.showerror("Erro", f"Falha ao inicializar planilha: {e}")
        return False

def verificar_e_inicializar_planilha():
    """
    Verifica e inicializa o sistema de planilhas.
    Deve ser chamada no início do programa.
    """
    if not os.path.exists(dados_planilha):
        # Mostra tela de loading enquanto inicializa
        loading_screen = LoadingScreen(root)
        loading_screen.update_progress(10, "Inicializando sistema...")
        
        try:
            if not carregar_ou_criar_planilha():
                messagebox.showerror("Erro", "Não foi possível inicializar a planilha de dados")
                root.destroy()
                return False
            
            loading_screen.update_progress(100, "Sistema pronto!")
            loading_screen.close()
            return True
            
        except Exception as e:
            loading_screen.close()
            messagebox.showerror("Erro", f"Falha na inicialização: {e}")
            root.destroy()
            return False
    return True


if not verificar_e_inicializar_planilha():
    sys.exit(1)

# Função para carregar coordenadas IPPUC na memória
def carregar_ippuc_na_memoria():
    global ippuc_cache
    try:
        if os.path.exists(coordenadas_ippuc_planilha):
            workbook = load_workbook(coordenadas_ippuc_planilha, read_only=True)
            sheet = workbook.active
            ippuc_cache = {
                str(row[0]): {  # ID IPPUC como chave
                    "latitude": row[1], 
                    "longitude": row[2],  
                }
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] is not None and row[1] is not None and row[2] is not None
            }
            logging.info(f"Dados de IPPUC carregados: {len(ippuc_cache)} registros.")
        else:
            logging.warning("Planilha de IPPUC não encontrada.")
    except Exception as e:
        logging.error(f"Erro ao carregar IPPUC: {e}")




# Função para carregar coordenadas na memória
def carregar_coordenadas_na_memoria():
    global coordenadas_cache
    try:
        if os.path.exists(coordenadas_planilha):
            workbook = load_workbook(coordenadas_planilha, read_only=True)
            sheet = workbook.active
            coordenadas_cache = {
                str(row[0]): {
                    "coordenadas": f"{row[7]},{row[8]}",
                    "latitude": row[7],
                    "longitude": row[8],
                    "bairro": row[5],
                    "distancia_postes": row[45],
                    "altura": row[39],
                    "projecao": row[38],
                    "recuo": row[43],
                }
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] is not None and row[7] is not None and row[8] is not None and row[45] is not None and row[39] is not None and row[38] is not None and row[43] is not None
            }
            logging.info("Coordenadas e bairros carregados na memória: %d registros.", len(coordenadas_cache))
        else:
            coordenadas_cache = {}
            logging.warning("Arquivo de coordenadas não encontrado.")
    except Exception as e:
        logging.error("Erro ao carregar coordenadas na memória: %s", e)
        coordenadas_cache = {}

# Função para buscar coordenadas
def buscar_coordenadas(id_raag, id_ippuc):
    try:
        # Prioridade 1: Busca por ID IPPUC
        if id_ippuc and id_ippuc.strip() in ippuc_cache:
            dados = ippuc_cache[id_ippuc.strip()]
            return (
                dados.get("latitude", ""),
                dados.get("longitude", ""),
                dados.get("bairro", ""),
                "",  # Distância entre postes (não está na planilha IPPUC)
                "",  # Altura (não está na planilha IPPUC)
                "",  # Projeção (não está na planilha IPPUC)
                ""   # Recuo (não está na planilha IPPUC)
            )
        
        # Prioridade 2: Busca por ID RAAG (lógica original)
        elif id_raag and id_raag.strip() in coordenadas_cache:
            dados = coordenadas_cache[id_raag.strip()]
            coords = dados.get("coordenadas", "").split(",")
            latitude = coords[0].strip() if len(coords) > 0 else ""
            longitude = coords[1].strip() if len(coords) > 1 else ""
            return (
                latitude,
                longitude,
                dados.get("bairro", ""),
                dados.get("distancia_postes", ""),
                dados.get("altura", ""),
                dados.get("projecao", ""),
                dados.get("recuo", "")
            )
        
        # Se não encontrar em nenhum:
        return "", "", "", "", "", "", ""
    
    except Exception as e:
        logging.error(f"Erro ao buscar coordenadas: {e}")
        return "", "", "", "", "", "", ""

# Função para preencher dados automaticamente
def preencher_dados_automaticamente(event=None):
    id_raag = entries["entry_id_raag"].get().strip()
    id_ippuc = entries["entry_id_ippuc"].get().strip()

    # Não faz nada se ambos os IDs estiverem vazios
    if not id_raag and not id_ippuc:
        return

    try:
        # Busca as coordenadas (priorizando IPPUC)
        latitude, longitude, bairro, distancia_postes, altura, projecao, recuo = buscar_coordenadas(id_raag, id_ippuc)

        # ===== PREENCHIMENTO DE CAMPOS BÁSICOS =====
        # Preenche latitude (se existir e o campo estiver vazio)
        if latitude:
            entries["entry_latitude"].delete(0, tk.END)
            entries["entry_latitude"].insert(0, str(latitude).strip())
        
        # Preenche longitude (se existir e o campo estiver vazio)
        if longitude:
            entries["entry_longitude"].delete(0, tk.END)
            entries["entry_longitude"].insert(0, str(longitude).strip())

        # Preenche bairro (se existir e o campo estiver vazio)
        if bairro and not entries["entry_bairro"].get().strip():
            entries["entry_bairro"].delete(0, tk.END)
            entries["entry_bairro"].insert(0, str(bairro).strip())

        # ===== CAMPOS ESPECÍFICOS DO RAAG (só preenche se não veio do IPPUC) =====
        if not id_ippuc or id_ippuc not in ippuc_cache:
            # ARREDONDAMENTO: Distância entre postes
            if distancia_postes and not entries["entry_distancia_postes"].get().strip():
                try:
                    distancia_str = str(distancia_postes).replace(",", ".")
                    distancia_float = float(distancia_str) if distancia_str else 0.0
                    distancia_arredondada = int(distancia_float + 0.5)
                    entries["entry_distancia_postes"].delete(0, tk.END)
                    entries["entry_distancia_postes"].insert(0, str(distancia_arredondada))
                except (ValueError, AttributeError):
                    pass

            # FORMATAÇÃO: Projeção
            if projecao and not entries["entry_projecao"].get().strip():
                try:
                    projecao_str = str(projecao).replace(",", ".")
                    projecao_float = float(projecao_str) if projecao_str else 0.0

                    if projecao_float == 0:
                        projecao_formatada = " "
                    elif projecao_float <= 1.2:
                        projecao_formatada = "1"
                    elif projecao_float == 1.9:
                        projecao_formatada = "1.5"
                    elif projecao_float <= 2.6:
                        projecao_formatada = "2.35"
                    else:
                        projecao_formatada = "AVALIAR"

                    entries["entry_projecao"].delete(0, tk.END)
                    entries["entry_projecao"].insert(0, projecao_formatada)
                except (ValueError, AttributeError):
                    pass

            # Preenche outros campos específicos do RAAG
            campos_raag = [
                ("entry_altura", altura),
                ("entry_recuo", recuo)
            ]

            for campo, valor in campos_raag:
                if valor and not entries[campo].get().strip():
                    entries[campo].delete(0, tk.END)
                    entries[campo].insert(0, str(valor).strip())

    except Exception as e:
        logging.error(f"Erro ao preencher dados automaticamente: {e}")
        messagebox.showerror(
            "Erro", 
            f"Falha ao carregar dados automáticos. Verifique os IDs ou valores de origem.\nDetalhes: {e}"
        )



# Função para carregar classificação na memória
def carregar_classificacao_na_memoria():
    global classificacao_cache
    try:
        if os.path.exists("Classificação.xlsx"):
            workbook = load_workbook("Classificação.xlsx", read_only=True)
            sheet = workbook.active
            classificacao_cache = {
                str(row[0]).strip().upper(): (row[1], row[2], row[4] if len(row) > 4 else "")  # Via: (Classe, Passeio, Trecho)
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] and row[1] and row[2]  # Garante que Via, Classe e Passeio existam
            }
            logging.info("Classificação e Trechos carregados na memória: %d registros.", len(classificacao_cache))
        else:
            classificacao_cache = {}
            logging.warning("Arquivo de classificação não encontrado.")
    except Exception as e:
        logging.error("Erro ao carregar classificação na memória: %s", e)
        classificacao_cache = {}

# Função para buscar classificação
def buscar_classificacao(via):
    try:
        via_normalizada = via.strip().upper()
        return classificacao_cache.get(via_normalizada, ("", "", ""))  # Retorna (Classe, Passeio, Trecho)
    except Exception as e:
        logging.error("Erro ao buscar classificação: %s", e)
        return "", "", ""
    
def carregar_dados_em_segundo_plano(loading_screen, callback):
    def tarefa():
        try:
            loading_screen.update_progress(10, "Carregando configurações...")
            
            # Primeiro verifica/cria a planilha
            if not os.path.exists(dados_planilha) and not os.path.exists(CONFIG_FILE):
                root.after(0, lambda: carregar_ou_criar_planilha())
                while not os.path.exists(dados_planilha):  # Aguarda criação
                    time.sleep(0.1)
            
            loading_screen.update_progress(30, "Carregando dados...")
            
            # Só carrega se a planilha existe
            if os.path.exists(dados_planilha):
                carregar_coordenadas_na_memoria()
                carregar_classificacao_na_memoria()
                carregar_ippuc_na_memoria()
            
            loading_screen.update_progress(90, "Preparando interface...")
            time.sleep(0.5)
            
            loading_screen.update_progress(100, "Pronto!")
            root.after(0, callback)
            
        except Exception as e:
            error_msg = str(e)
            root.after(0, lambda: loading_screen.close())  # Fecha loading antes de mostrar erro
            root.after(0, lambda: messagebox.showerror("Erro", f"Falha no carregamento: {error_msg}"))
            root.after(0, lambda: root.destroy() if not os.path.exists(dados_planilha) else callback())

    threading.Thread(target=tarefa, daemon=True).start()
  

def preencher_classificacao(event):
    via = entries["entry_via"].get().strip()
    if via:
        via_normalizada = via.upper()
        classe_via, passeio, trecho = buscar_classificacao(via)
        
        # Preenche Classificação (se vazio)
        if classe_via and passeio and not entries["entry_classificacao"].get().strip():
            classificacao_completa = f"{classe_via}-{passeio}"
            entries["entry_classificacao"].delete(0, tk.END)
            entries["entry_classificacao"].insert(0, classificacao_completa)
        
        # Preenche Trecho (se vazio)
        if trecho and not entries["entry_trecho"].get().strip():
            entries["entry_trecho"].delete(0, tk.END)
            entries["entry_trecho"].insert(0, trecho)


# Função para salvar dados
def salvar_dados():
    if not validar_campos():
        return

    # Executa a função de salvar em segundo plano
    threading.Thread(target=salvar_em_segundo_plano, daemon=True).start()


# Função para exibir a mensagem temporária
def exibir_mensagem_temporaria(mensagem):
    label_mensagem.config(text=mensagem)  # Exibe a mensagem
    label_mensagem.grid()  # Torna o Label visível
    root.after(4000, lambda: label_mensagem.grid_remove())  # Oculta após 4 segundos

# Modifique a função salvar_em_segundo_plano
def salvar_em_segundo_plano():
    try:
        # Verifica se a planilha existe e está acessível
        if not verificar_e_inicializar_planilha():
            # Tenta criar a planilha se não existir
            if not carregar_ou_criar_planilha():
                messagebox.showerror("Erro", "Planilha não encontrada e não pôde ser criada")
                return
        
        # Carrega a planilha (sem usar context manager)
        workbook = load_workbook(dados_planilha)
        sheet = workbook["Levantamento"]

        # Coleta os dados do formulário
        dados = {key: entry.get().strip() for key, entry in entries.items()}

        id_raag = dados["entry_id_raag"]
        linha_existente = None

        # Verifica se o ID RAAG já existe na planilha
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == id_raag:
                linha_existente = row[0].row
                break

        # Verifica se o ID RAAG já existe em outro registro (apenas para novos registros)
        if not linha_existente:
            for row in sheet.iter_rows(min_row=2):
                if str(row[0].value) == id_raag:
                    messagebox.showerror("Erro", "Este ID RAAG já existe na planilha.")
                    workbook.close()  # Fecha o workbook antes de retornar
                    return

        # Campos que devem manter os valores atuais caso não sejam preenchidos novamente
        campos_preservados = {
            "entry_largura_passeio_adj", "entry_largura_gramado_adj", "entry_largura_estac_adj",
            "entry_largura_pista1", "entry_largura_canteiro_central", "entry_largura_pista2",
            "entry_largura_estac_opo", "entry_largura_gramado_opo", "entry_largura_passeio_opo",
            "entry_ciclovia", "entry_via", "entry_bairro", "entry_trecho", "entry_classificacao",  # Correção: vírgula adicionada aqui
            "combobox_distribuicao",
        }

        if linha_existente:
            # Editar registro existente
            valores_atuais = {}  # Armazena valores atuais antes da edição
            for idx, key in enumerate(entries.keys(), start=1):
                valores_atuais[key] = sheet.cell(row=linha_existente, column=idx).value

            for idx, (chave, valor) in enumerate(dados.items(), start=1):
                if chave == "entry_id_raag" and valor.strip() != str(sheet.cell(row=linha_existente, column=1).value):
                    # Verifica se o novo ID RAAG é diferente do atual na planilha e se já existe em outra linha
                    if any(str(cell.value) == valor for row in sheet.iter_rows(min_row=2) for cell in row if cell.row != linha_existente):
                        root.after(0, messagebox.showerror, "Erro", "Este ID RAAG já existe na planilha.")
                        workbook.close()  # Fecha o workbook antes de retornar
                        return  # Impede a edição se o novo ID RAAG já existir
                    else:
                        sheet.cell(row=linha_existente, column=idx, value=valor)  # Permite a alteração do ID RAAG
                elif chave in campos_preservados and not valor.strip():
                    # Mantém o valor atual para os outros campos preservados se estiverem vazios
                    sheet.cell(row=linha_existente, column=idx, value=valores_atuais[chave])
                else:
                    # Atualiza outros campos normalmente
                    sheet.cell(row=linha_existente, column=idx, value=valor)

            root.after(0, exibir_mensagem_temporaria, f"Dados do ID RAAG {id_raag} editados com sucesso!")
        
        else:  # Inserir novo registro
            prox_linha = sheet.max_row + 1
            for idx, valor in enumerate(dados.values(), start=1):
                sheet.cell(row=prox_linha, column=idx, value=valor)
            root.after(0, exibir_mensagem_temporaria, "Dados salvos com sucesso!")

        # Salva as alterações e fecha o workbook
        workbook.save(dados_planilha)
        workbook.close()
        
        logging.info("Dados salvos/editados: %s", dados)

        # Limpa os campos (exceto os preservados)
        for key, entry in entries.items():
            if key not in campos_preservados:
                if isinstance(entry, ttk.Combobox):
                    entry.set("Selecione")
                else:
                    entry.delete(0, tk.END)

        atualizar_treeview()

    except Exception as e:
        logging.error("Erro ao salvar dados: %s", e)
        root.after(0, messagebox.showerror, "Erro", f"Ocorreu um erro ao salvar os dados: {e}")
        # Garante que o workbook seja fechado mesmo em caso de erro
        if 'workbook' in locals():
            workbook.close()

def exportar_para_csv():
    try:
        # Carrega a planilha
        workbook = load_workbook(dados_planilha)
        sheet = workbook["Levantamento"]

        # Nome do arquivo CSV de saída
        nome_csv = "coordenadas_excel.csv"
        
        # Cria o arquivo CSV com delimitador ";"
        with open(nome_csv, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file, delimiter=';')
            
            # Cabeçalho (separado por ";")
            writer.writerow(["ID_RAAG", "ID_IPPUC", "LATITUDE", "LONGITUDE"])
            
            # Escreve os dados
            for row in sheet.iter_rows(min_row=2, values_only=True):
                id_raag = row[0] if row[0] is not None else ""
                id_ippuc = row[1] if row[1] is not None else ""
                latitude = row[7] if row[7] is not None else ""
                longitude = row[8] if row[8] is not None else ""
                
                writer.writerow([id_raag, id_ippuc, latitude, longitude])

        messagebox.showinfo(
            "Sucesso", 
            f"Arquivo '{nome_csv}' gerado com sucesso!\n"
            "O Excel deve abrir as colunas automaticamente."
        )
    
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao exportar: {str(e)}")

def atualizar_treeview():
    global vias
    vias = carregar_vias_e_ids()
    for item in tree_vias.get_children():
        tree_vias.delete(item)

    for via, ids in vias.items():
        # Converte os IDs para strings antes de inseri-los na Treeview
        ids_str = [str(id_raag) for id_raag in ids]  # ou tuple(map(str, ids))
        tree_vias.insert("", "end", text=via, values=ids_str)

    filtrar_vias()


def avancar_campo(event):
    # Obtém o widget que recebeu o evento
    widget_atual = event.widget

    # Obtém a aba atual
    aba_atual = notebook.index(notebook.select())

    # Lista de campos na aba atual
    campos_aba = []
    for child in notebook.winfo_children()[aba_atual].winfo_children():
        if isinstance(child, (ttk.Entry, ttk.Combobox)):
            campos_aba.append(child)

    # Encontra o índice do campo atual na lista de campos da aba
    try:
        indice_atual = campos_aba.index(widget_atual)
    except ValueError:
        return  # Se o campo não estiver na lista, não faz nada

    # Avança para o próximo campo
    if indice_atual + 1 < len(campos_aba):
        proximo_campo = campos_aba[indice_atual + 1]
        proximo_campo.focus_set()  # Foca no próximo campo
    else:
        # Se não houver mais campos na aba, muda para a próxima aba
        if aba_atual + 1 < len(notebook.tabs()):
            notebook.select(aba_atual + 1)  # Muda para a próxima aba
            # Foca no primeiro campo da próxima aba
            primeiro_campo_proxima_aba = None
            for child in notebook.winfo_children()[aba_atual + 1].winfo_children():
                if isinstance(child, (ttk.Entry, ttk.Combobox)):
                    primeiro_campo_proxima_aba = child
                    break
            if primeiro_campo_proxima_aba:
                primeiro_campo_proxima_aba.focus_set()
        else:
            # Se for o último campo da última aba, salva os dados
            salvar_dados()


# Função para validar campos
def validar_campos():
    # Campos que não são obrigatórios
    campos_nao_obrigatorios = {
        "entry_id_ippuc",  # Já não era obrigatório
        "entry_modelo_led",  # Novo campo não obrigatório
        "combobox_observacoes",  # Novo campo não obrigatório
        "entry_observacoes_gerais",  # Novo campo não obrigatório
        "entry_trecho",  # Novo campo não obrigatório
    }

    # Verifica campos obrigatórios
    for key, entry in entries.items():
        if key in campos_nao_obrigatorios:  # Ignora campos não obrigatórios
            continue
        if isinstance(entry, ttk.Combobox) and entry.get() == "Selecione":
            messagebox.showerror("Erro", f"O campo '{key}' deve ser preenchido.")
            return False
        elif isinstance(entry, tk.Entry) and not entry.get().strip():
            messagebox.showerror("Erro", f"O campo '{key}' deve ser preenchido.")
            return False

    # ===== NOVA VALIDAÇÃO: Projeção não pode ser "AVALIAR" =====
    projecao = entries["entry_projecao"].get().strip()
    if projecao.upper() == "AVALIAR":
        messagebox.showerror(
            "Erro",
            "A projeção está como 'AVALIAR'. Por favor, ajuste o valor antes de salvar."
        )
        return False

    return True


# Função para limpar campos
def limpar_campos():
    resposta = messagebox.askyesno("Confirmação", "Tem certeza de que deseja limpar todos os campos?")
    if resposta:  # Apenas limpa os campos se a resposta for "Sim"
        for entry in entries.values():
            if isinstance(entry, ttk.Combobox):
                entry.set("Selecione")
            else:
                entry.delete(0, tk.END)

def deletar_id_raag(event=None):
    item_selecionado = tree_vias.selection()
    if item_selecionado:
        item = item_selecionado[0]
        valores = tree_vias.item(item, "values")
        if valores:  # Verifica se é um ID RAAG (tem valores)
            id_raag = valores[0]
            confirmacao = messagebox.askyesno("Confirmar Exclusão", f"Tem certeza que deseja excluir o ID RAAG {id_raag}?")
            if confirmacao:
                try:
                    workbook = load_workbook(dados_planilha)
                    sheet = workbook["Levantamento"]

                    linha_para_deletar = None
                    for row in sheet.iter_rows(min_row=2):
                        if str(row[0].value) == id_raag:
                            linha_para_deletar = row[0].row
                            break

                    if linha_para_deletar:
                        sheet.delete_rows(linha_para_deletar)
                        workbook.save(dados_planilha)

                        # Remove o ID RAAG da Treeview
                        tree_vias.delete(item)
                        atualizar_treeview()  # Atualiza a treeview após a exclusão

                        messagebox.showinfo("Sucesso", f"ID RAAG {id_raag} excluído com sucesso!")
                        limpar_campos()

                    else:
                        messagebox.showwarning("Aviso", f"ID RAAG {id_raag} não encontrado na planilha.")

                except Exception as e:
                    logging.error("Erro ao excluir ID RAAG: %s", e)
                    messagebox.showerror("Erro", f"Ocorreu um erro ao excluir o ID RAAG: {e}")


# Função para mudar de aba
def mudar_aba(direcao):
    """
    Muda para a aba anterior ou próxima.
    :param direcao: "anterior" ou "proxima"
    """
    atual = notebook.index(notebook.select())  # Obtém o índice da aba atual
    if direcao == "anterior" and atual > 0:
        notebook.select(atual - 1)  # Muda para a aba anterior
    elif direcao == "proxima" and atual < len(notebook.tabs()) - 1:
        notebook.select(atual + 1)  # Muda para a próxima aba

# Função para carregar vias e IDs RAAG
def carregar_vias_e_ids():
    try:
        if not os.path.exists(dados_planilha):
            return {}  # Retorna um dicionário vazio se o arquivo não existir
            
        workbook = load_workbook(dados_planilha)
        sheet = workbook["Levantamento"]
        vias = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 2:  # Verifica se a linha existe e tem pelo menos 3 colunas
                via = row[2] if row[2] else "Sem nome"
                id_raag = str(row[0]) if row[0] else "Sem ID"
                if via not in vias:
                    vias[via] = []
                vias[via].append(id_raag)
                
        return vias
        
    except Exception as e:
        logging.error(f"Erro ao carregar vias e IDs RAAG: {e}")
        return {}  # Retorna um dicionário vazio em caso de erro

# Função para filtrar vias
def filtrar_vias(event=None):
    termo = entry_busca_via.get().strip().lower()
    for via in tree_vias.get_children():
        tree_vias.delete(via)
    for via, ids in vias.items():
        if termo in via.lower():
            tree_vias.insert("", "end", text=via, values=ids)

# Função para carregar dados por ID RAAG
def carregar_dados_por_id(id_raag):
    try:
        workbook = load_workbook(dados_planilha)
        sheet = workbook["Levantamento"]
        dados_encontrados = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == id_raag:
                dados_encontrados = row
                break

        if dados_encontrados:
            for idx, valor in enumerate(dados_encontrados):
                coluna = list(entries.keys())[idx]
                if valor:  # Apenas insere se houver valor
                    if isinstance(entries[coluna], ttk.Combobox):
                        entries[coluna].set(valor)
                    else:
                        entries[coluna].delete(0, tk.END)
                        entries[coluna].insert(0, valor)
            messagebox.showinfo("Sucesso", f"Dados do ID RAAG {id_raag} carregados com sucesso!")
        else:
            messagebox.showwarning("Aviso", f"ID RAAG {id_raag} não encontrado.")

    except Exception as e:
        logging.error("Erro ao carregar dados: %s", e)
        messagebox.showerror("Erro", "Ocorreu um erro ao carregar os dados.")

def selecionar_via_ou_id(event):
    item = tree_vias.selection()[0]
    texto = tree_vias.item(item, "text")
    valores = tree_vias.item(item, "values")

    if texto:  # Verifica se o texto não está vazio (é uma via)
        if tree_vias.item(item, 'open'):
            tree_vias.item(item, open=False)  # Recolhe a via se estiver aberta
        else:
            if texto in vias:
                # Verifica se a via já tem filhos (IDs RAAG)
                if not tree_vias.get_children(item):  # Insere somente se não houver filhos
                    for id_raag in vias[texto]:
                        tree_vias.insert(item, "end", text="", values=(str(id_raag),))
                tree_vias.item(item, open=True)  # Expande a via

    elif valores:  # Se não for uma via e tiver valores, é um ID RAAG
        id_raag = valores[0]
        carregar_dados_por_id(id_raag)

        # Expande a via pai se não estiver expandida (mesma lógica anterior)
        pai = tree_vias.parent(item)
        if pai and not tree_vias.item(pai, 'open'):
            tree_vias.item(pai, open=True)

# Função para recolher/expandir a seção lateral
def toggle_frame_vias():
    if frame_vias.winfo_ismapped():
        frame_vias.grid_remove()  # Use grid_remove()
        botao_toggle.config(text="Mostrar Vias")
    else:
        frame_vias.grid()      # Use grid()
        botao_toggle.config(text="Recolher Vias")
        root.update_idletasks() # Force layout update


# Cria e mostra a tela de loading
loading_screen = LoadingScreen(root)


# Frame fixo à esquerda para vias e busca
frame_vias = ttk.Frame(root)
frame_vias.grid(row=0, column=0, sticky="ns", padx=10, pady=10)

# Conteúdo do frame_vias
label_busca_via = ttk.Label(frame_vias, text="Buscar Via:")
label_busca_via.pack(pady=5)

entry_busca_via = ttk.Entry(frame_vias)
entry_busca_via.pack(pady=5, fill="x")
entry_busca_via.bind("<KeyRelease>", filtrar_vias)

# Configuração da Treeview para exibir vias e IDs RAAG
tree_vias = ttk.Treeview(frame_vias, columns=("ids_raag",), show="tree headings")
tree_vias.heading("#0", text="Via")
tree_vias.heading("ids_raag", text="IDs RAAG")
tree_vias.column("ids_raag", width=150)  # Define a largura da coluna
tree_vias.pack(fill="both", expand=True)
tree_vias.bind("<Double-1>", selecionar_via_ou_id)

# Carregar vias e IDs RAAG
vias = carregar_vias_e_ids()
for via, ids in vias.items():
    tree_vias.insert("", "end", text=via)

# Interface gráfica principal (Corrected to use grid)
notebook = ttk.Notebook(root)
notebook.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

# Associar eventos de teclado para mudar de aba
root.bind("<Left>", lambda event: mudar_aba("anterior"))  # Seta para a esquerda
root.bind("<Right>", lambda event: mudar_aba("proxima"))  # Seta para a direita

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)

aba1 = ttk.Frame(notebook)
aba2 = ttk.Frame(notebook)
aba3 = ttk.Frame(notebook)
notebook.add(aba1, text="Dados Gerais")
notebook.add(aba2, text="Larguras e Ciclovia")
notebook.add(aba3, text="Postes e Interferências")

for aba in [aba1, aba2, aba3]:
    aba.grid_columnconfigure(0, weight=1)
    aba.grid_columnconfigure(1, weight=2)

def criar_campos(aba, fields):
    for idx, (label_text, field_name) in enumerate(fields, start=1):
        label = ttk.Label(aba, text=label_text)
        label.grid(row=idx, column=0, padx=5, pady=5, sticky=tk.W)

        if "combobox" in field_name.lower():
            values = ["Selecione"]
            if field_name == "combobox_distribuicao":
                values.extend(["UNILATERAL", "BILATERAL ALTERNADO", "BILATERAL FF"])
            elif field_name == "combobox_interferencia_arborea":
                values.extend(["SIM", "NÃO"])
            elif field_name == "combobox_led":
                values.extend(["SIM", "NÃO"])
            elif field_name == "combobox_observacoes":
                values.extend(["Retirar braço com luminária", "Instalar braço com luminária", "Instalar braço, luminária e suporte p/ poste duplo T", "Instalar braço, luminária e suporte p/ poste circular",  "Rotacionar braço para a via", "Braço muito próximo do Transformador", 
                "Braço com luminária virado para propriedade privada", "iluminação exclusiva para passeio", "Poste com duas luminárias, com uma virada para calçada",
                 "Medidas imprecisas, validar", "Não é possivel levantar as informações", "Ajustar inclinação do braço", "Instalar POSTE, braço e luminária"])

            combobox = ttk.Combobox(aba, values=values, state="readonly")
            combobox.set("Selecione")
            combobox.grid(row=idx, column=1, padx=5, pady=5, sticky="ew")
            combobox.bind("<Return>", avancar_campo)  # Adiciona evento Enter
            entries[field_name] = combobox
        else:
            entry = ttk.Entry(aba)
            entry.grid(row=idx, column=1, padx=5, pady=5, sticky="ew")
            entry.bind("<Return>", avancar_campo)  # Adiciona evento Enter
            entries[field_name] = entry

# Dicionário para armazenar os campos
entries = {}

# Criar campos na aba 1
criar_campos(aba1, [
    ("ID RAAG", "entry_id_raag"),
    ("ID IPPUC", "entry_id_ippuc"),
    ("Via", "entry_via"),
    ("Bairro", "entry_bairro"),
    ("Trecho", "entry_trecho"),
    ("Classificação", "entry_classificacao"),
    ("Distribuição", "combobox_distribuicao"),
    ("Latitude", "entry_latitude"),
    ("Longitude", "entry_longitude"),
])

# Associar evento para preenchimento automático da classificação
entries["entry_via"].bind("<FocusOut>", preencher_classificacao)

# Criar campos na aba 2
criar_campos(aba2, [
    ("Largura do Passeio Adjacente", "entry_largura_passeio_adj"),
    ("Largura do Gramado Adjacente", "entry_largura_gramado_adj"),
    ("Largura do Estacionamento Adjacente", "entry_largura_estac_adj"),
    ("Largura da Pista 1", "entry_largura_pista1"),
    ("Largura do Canteiro Central", "entry_largura_canteiro_central"),
    ("Largura da Pista 2", "entry_largura_pista2"),
    ("Largura do Estacionamento Oposto", "entry_largura_estac_opo"),
    ("Largura do Gramado Oposto", "entry_largura_gramado_opo"),
    ("Largura do Passeio Oposto", "entry_largura_passeio_opo"),
    ("Ciclovia", "entry_ciclovia"),
])

# Criar campos na aba 3
criar_campos(aba3, [
    ("Distância entre Postes", "entry_distancia_postes"),
    ("Altura", "entry_altura"),
    ("Projeção", "entry_projecao"),
    ("Recuo", "entry_recuo"),
    ("Interferência Arbórea", "combobox_interferencia_arborea"),
    ("Led", "combobox_led"),
    ("Modelo do LED", "entry_modelo_led"),  # Novo campo
    ("Observações", "combobox_observacoes"),  # Novo campo (Combobox)
    ("Observações Gerais", "entry_observacoes_gerais"),  # Novo campo
])



# Adicione um Label para a mensagem temporária
label_mensagem = ttk.Label(root, text="", foreground="green")
label_mensagem.grid(row=3, column=1, sticky="ew", padx=10, pady=10)

# Botões principais (incluindo o botão de alternar vias)
frame_botoes = ttk.Frame(root)
frame_botoes.grid(row=2, column=1, sticky="ew", padx=10, pady=10)

botao_salvar = ttk.Button(frame_botoes, text="Salvar", command=salvar_dados)
botao_salvar.grid(row=0, column=0, padx=5, pady=5)

botao_limpar = ttk.Button(frame_botoes, text="Limpar Campos", command=limpar_campos)
botao_limpar.grid(row=0, column=1, padx=5, pady=5)

# Adicione este botão junto aos outros (no frame_botoes)
botao_exportar = ttk.Button(
    frame_botoes, 
    text="Exportar para CSV", 
    command=exportar_para_csv
)
botao_exportar.grid(row=0, column=3, padx=5, pady=5)  # Ajuste a posição conforme necessário

botao_toggle = ttk.Button(frame_botoes, text="Recolher Vias", command=toggle_frame_vias)
botao_toggle.grid(row=0, column=2, padx=5, pady=5)  # Ajuste a posição do botão "Recolher Vias" 

tree_vias.bind("<Double-1>", selecionar_via_ou_id)

entries["entry_id_ippuc"].bind("<FocusOut>", preencher_dados_automaticamente)
entries["entry_id_raag"].bind("<FocusOut>", preencher_dados_automaticamente)

atualizar_treeview()  # Inicializa a Treeview


tree_vias.bind("<Delete>", deletar_id_raag)

# Função chamada quando o carregamento terminar
def carregamento_concluido():
    try:
        loading_screen.close()
        root.deiconify()
        root.update_idletasks()
        root.geometry("1000x600")
        
        if os.path.exists(dados_planilha):
            atualizar_treeview()
        else:
            messagebox.showinfo("Informação", "Planilha não encontrada. Você pode criar uma nova ao salvar o primeiro registro.")
            
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao inicializar interface: {str(e)}")
        root.destroy()

# Inicia o carregamento
carregar_dados_em_segundo_plano(loading_screen, carregamento_concluido)

root.mainloop()